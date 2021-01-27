from PIL import Image
from openpyxl import load_workbook

img = Image.open('s34hunka.jfif')
pixels = img.load()
width, height = img.size

wb = load_workbook('s34hunka.xlsx')
ws = wb.active

for j in range(height):
    for i in range(width):
        c = ws.cell(j + 1,i + 1)
        p = pixels[i,j]
        color = ''.join('%02X'%t for t in p)
        if color == c.fill.fgColor.rgb[2:]:
            pixels[i,j] = (255, 255, 255)
        else:
            pixels[i,j] = (0, 0, 0)
            
img.save("myresult.png")